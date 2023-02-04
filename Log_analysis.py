# -*- coding: UTF-8 -*-
import PyQt5
from PyQt5 import QtWidgets, QtCore
from PyQt5.QtWidgets import QApplication, QMainWindow, QMenuBar, QMenu, QFileDialog, QCheckBox
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import PatternFill
import threading
import toml
import os
import pandas as pd
import csv
import os.path
import re
import sys
import time
import openpyxl
from openpyxl.worksheet.table import Table



def read_toml(file_path):
    with open(file_path, "r") as f:
        return toml.load(f)
config = read_toml('Config_file.ini')

# ------ Входящий поток -----------
name_a_shared_csv = config['Name_a_shared_csv'] # название общего csv файла
CPU_load = config['CPU_load']
User_byte = config['User_bytes']
#-----------------------------------

#------- Исходящий поток -----------
flag_sender_stat = config['sender_stat']
fv_input = config['fv_input']
fa_input = config['fa_input']
iv_input = config['iv_input']
ia_input = config['ia_input']
#-----------------------------------


def creating_a_shared_csv(name_folder):
    ''' Функция создания общего файла по пользователям'''

    if os.path.exists(f'{name_folder}\{name_a_shared_csv}.xlsx') != True:
        files = os.listdir(path=f"{name_folder}")
        with open(f'{name_folder}\{name_a_shared_csv}.csv', 'w', newline='') as file:
            header = [(
                'Дата',
                'Время',
                'ID пользователя',
                'Суть проблемы',
                'Значение проблемного показателя'
            )]
            writer = csv.writer(file, delimiter=';')
            writer.writerows(header)
            i = 0
            while i < len(files):
                if files[i] != f'{name_a_shared_csv}.csv' and files[i] != 'load_CPU_users.xlsx' and files[i] != 'load_CPU_users.csv' and ('sender_stat' not in files[i]):

                    anime = pd.read_excel(f'{name_folder}\{files[i]}', skiprows=[0], header=None)


                    r = anime[anime[9] >= 0]
                    r = r.sort_values(by=[9], ascending=False).head(1)

                    zn_bytes = int(r[9][:1])
                    if zn_bytes > User_byte:
                        name_user = r[2][:1]
                        name_user = list(name_user)[0].split('@')[0]
                        message = 'Проблемы с каналом связи'

                        date, time, users_butes = r[0][:1].item(), r[1][:1].item(), str(r[9][:1].item()) + '  кол-во байт в очереди'
                        frame = [date, time, name_user, message, users_butes]

                        writer = csv.writer(file, delimiter=';')
                        writer.writerow(frame)


                    CPU = anime[anime[3] >= 0]
                    CPU = CPU.sort_values(by=[3], ascending=False).head(1)

                    load_CPU = int(CPU[3][:1])
                    if load_CPU > CPU_load:
                        name_user = CPU[2][:1]
                        name_user = list(name_user)[0].split('@')[0]
                        message = 'Загрузка ЦП'

                        date, time, load = CPU[0][:1].item(), CPU[1][:1].item(), str(CPU[3][:1].item()) + '%' + '  загрузка ЦП'
                        frame = [date, time, name_user, message, load]

                        writer = csv.writer(file, delimiter=';')
                        writer.writerow(frame)

                i += 1

        # -------------- конвертация файла с csv формата в xlsx --------------------------------------
        read_file = pd.read_csv(f'{name_folder}\{name_a_shared_csv}.csv', sep=';', encoding='windows-1251')  # считываем содержимое csv файла
        read_file.to_excel(f'{name_folder}\{name_a_shared_csv}.xlsx', index=None, header=True)  # записываем содержимое в xlsx файл
        os.remove(f'{name_folder}\{name_a_shared_csv}.csv')  # удаляем csv файл
        # --------------------------------------------------------------------------------------------

        book = openpyxl.load_workbook(f'{name_folder}/{name_a_shared_csv}.xlsx')
        sheet = book.active
        sheet.column_dimensions["A"].width = 12
        sheet.column_dimensions["B"].width = 12
        sheet.column_dimensions["C"].width = 30
        sheet.column_dimensions["D"].width = 28
        sheet.column_dimensions["E"].width = 38
        fill1 = PatternFill('solid', fgColor='00cccc')
        sheet["A1"].fill = fill1
        sheet["B1"].fill = fill1
        sheet["C1"].fill = fill1
        sheet["D1"].fill = fill1
        sheet["E1"].fill = fill1

        nb_row = sheet.max_row  # узнаём колиечтво строк в эксель файле


        table = Table(displayName="Sheet1", ref=f'A1:E{nb_row}')  # создаем объект таблицы
        sheet.add_table(table)  # добавляем таблицу

        book.save(f'{name_folder}/{name_a_shared_csv}.xlsx')

#--------------------------------------------------------------------------------------------------------------
def create_user_csv(ID_user, name_folder, folder_path_log):
    '''Функция создания файла со статистикой входящего потока для каждого пользователя'''
    #name_user = ID_user.split('@')[0]
    #print('Name_user:::', name_user) # 1
    #folder_path = os.getcwd()
    #print('Name_folder:', name_folder) # vs_stat_svc_000000a06aab935d@ua7dv.trueconf.name#vcs
    #print('Folder_path:', folder_path) # C:\Users\marinich\PycharmProjects\Marinich  !!! пока не использую
    #print('Folder_path_log:', folder_path_log) # C:/TrueConf/svc_logs/vs_stat_svc_000000a06aab935d@ua7dv.trueconf.name#vcs.txt

    if '/' in ID_user:
        name_user = ID_user.replace(':', '').split('/')[0].split('@')[0]

    elif '!' in ID_user:
        name_user = ID_user.replace(':', '').split('!')[1]
        name_user = name_user.split('@')[0] if '@' in name_user else name_user

    else:
        name_user = ID_user.replace(':', '').split('@')[0]


    if os.path.exists(f'{name_folder}/{name_user}.xlsx') != True:
        with open(f'{name_folder}\{name_user}.csv', 'w', newline='') as file:
            header = [(
                'Дата',
                'Время',
                'ID пользователя',
                'Загрузка ЦП',
                'rcv',
                'audio',
                'Bndph',
                'Bndes',
                'Pkt',
                'Bytes',
                'vpart'
            )]
            writer = csv.writer(file, delimiter=';')
            writer.writerows(header)
            with open(f'{folder_path_log}', 'r') as f: # r'C:\Users\marinich\Desktop\vs_stat_svc_00000065214657dc@ua7dv.trueconf.name#vcs.txt'
                while True:
                    row_data = f.readline()
                    ch = r'(\)'.split('(')[1].split(')')[0]
                    row_data = row_data.replace(ch, '!')
                    if not row_data:
                        break
                    patter = r"(\d{2}/\d{2}/\d{4})\s(\d{2}:\d{2}:\d{2})\|\s+"f'({ID_user})'r"\s\:\s+\d+\,\s+(\d+)\s\|\s+(\d+)\s+\S\s+(\d+)\,\s+\d+\S\,\s+(\d+)\,\s+(\d+)\,\s+(\d+)\,\s+(\d+)\|\s+(\d+\s+\S\s+\d+)\,"
                    line_user = re.findall(patter, row_data)
                    if line_user != []:
                        date, time, load, rcv, audio, bndhp, bndes, pkt, users_butes, vpart = line_user[0][0], line_user[0][1], line_user[0][3], line_user[0][4], line_user[0][5], line_user[0][6], line_user[0][7], line_user[0][8], line_user[0][9], line_user[0][10]

                        lst = [date, time, name_user, load, rcv, audio, bndhp, bndes, pkt, users_butes, vpart]

                        writer = csv.writer(file, delimiter=';') # ,lineterminator="\r"
                        writer.writerow(lst)

        # -------------- конвертация файла name_user.csv с csv формата в xlsx  --------------------------------------
        read_file = pd.read_csv(f'{name_folder}\{name_user}.csv', sep=';', encoding='windows-1251')  # считываем содержимое csv файла
        read_file.to_excel(f'{name_folder}\{name_user}.xlsx', index=None, header=True)  # записываем содержимое в xlsx файл
        os.remove(f'{name_folder}\{name_user}.csv')  # удаляем csv файл
        # ----- Стиль файла name_user.xlsx ------

        book = openpyxl.load_workbook(f'{name_folder}/{name_user}.xlsx')
        sheet = book.active
        sheet.column_dimensions["A"].width = 12
        sheet.column_dimensions["B"].width = 12
        sheet.column_dimensions["C"].width = 30
        sheet.column_dimensions["D"].width = 17
        sheet.column_dimensions["E"].width = 11
        sheet.column_dimensions["F"].width = 12
        sheet.column_dimensions["G"].width = 12
        sheet.column_dimensions["H"].width = 12
        sheet.column_dimensions["I"].width = 11
        sheet.column_dimensions["J"].width = 12
        sheet.column_dimensions["K"].width = 12
        fill1 = PatternFill('solid', fgColor='00cccc')
        sheet["A1"].fill = fill1
        sheet["B1"].fill = fill1
        sheet["C1"].fill = fill1
        sheet["D1"].fill = fill1
        sheet["E1"].fill = fill1
        sheet["F1"].fill = fill1
        sheet["G1"].fill = fill1
        sheet["H1"].fill = fill1
        sheet["I1"].fill = fill1
        sheet["J1"].fill = fill1
        sheet["K1"].fill = fill1

        nb_row = sheet.max_row  # узнаём колиечтво строк в эксель файле
        table = Table(displayName="Sheet1", ref=f'A1:K{nb_row}')  # создаем объект таблицы
        sheet.add_table(table)  # добавляем таблицу

        #--------------------------------------------------------------------------
        book.create_sheet("Chart")
        sheet2 = book["Chart"]

        chart = LineChart()
        chart.title = "Количество неотправленных байт в очереди"
        chart.y_axis.title = "byte"
                                                    # График 1
        chart.x_axis.number_format = 'H-M'
        chart.x_axis.majorTimeUnit = "time"
        chart.x_axis.title = "Время"

        chart.add_data(f'Sheet1!J2:J{nb_row}')
        time = Reference(sheet, min_col=2, min_row=2, max_row=nb_row)
        chart.set_categories(time)
        chart.width = 30
        chart.height = 10

        sheet2.add_chart(chart, "A1")
        #-------------------------------------------
        chart2 = LineChart()
        chart2.title = "Физическая пропускная способность канала связи"
        chart2.y_axis.title = "Kbit/s"
                                                    # График 2
        chart2.x_axis.number_format = 'H-M'
        chart2.x_axis.majorTimeUnit = "time"
        chart2.x_axis.title = "Время"
        chart2.add_data(f'Sheet1!G2:G{nb_row}')

        time = Reference(sheet, min_col=2, min_row=2, max_row=nb_row)
        chart2.set_categories(time)
        chart2.width = 30
        chart2.height = 10

        sheet2.add_chart(chart2, "A22")
        #-------------------------------------------
        chart3 = LineChart()
        chart3.title = "Загрузка ЦП"
        chart3.y_axis.title = "%"
                                                    # График 3
        chart3.x_axis.number_format = 'H-M'
        chart3.x_axis.majorTimeUnit = "time"
        chart3.x_axis.title = "Время"


        chart3.add_data(f'Sheet1!D2:D{nb_row}')
        time = Reference(sheet, min_col=2, min_row=2, max_row=nb_row)
        chart3.set_categories(time)
        chart3.width = 30
        chart3.height = 10

        sheet2.add_chart(chart3, "A44")
        #------------------------------------------

        # --------------- Лист и столбец для ТП -------------------------------------------------------
        book.create_sheet("Для_ТП")
        sheet3 = book["Для_ТП"]
        sheet3.column_dimensions["A"].width = 20
        sheet3["A1"].fill = fill1
        sheet3["A1"] = "Опорное значение"
        for i in range(2, nb_row + 1):
            sheet3.cell(row=i, column=1, value=0)

        table = Table(displayName="Для_ТП", ref=f'A1:A{nb_row}')
        sheet3.add_table(table)
        chart2.add_data(f'Для_ТП!A2:A{nb_row}') # добавляем вторую линию на второй график
        # ---------------------------------------------------------------------------------------------



        book.save(f'{name_folder}/{name_user}.xlsx') # сохраняем и закрываем файл
        # --------------------------------------------------------------------------------------------

def load_users_csv(ID_user, name_folder, date, time, load):
    ''' Фунцкия создания файла с загрузкой ЦП пользователей конференции'''

    if '/' in ID_user:
        name_user = ID_user.replace(':', '').split('/')[0].split('@')[0]

    elif '!' in ID_user:
        name_user = ID_user.replace(':', '').split('!')[1]
        name_user = name_user.split('@')[0] if '@' in name_user else name_user

    else:
        name_user = ID_user.replace(':', '').split('@')[0]

    if os.path.exists(f'{name_folder}\load_CPU_users.csv') != True:
        with open(f'{name_folder}\load_CPU_users.csv', 'w', newline='') as file:
            header = [(
                'Дата',
                'Время',
                'ID пользователя',
                'Загрузка ЦП',
                'Суть проблемы'
            )]
            writer = csv.writer(file, delimiter=';')
            writer.writerows(header)

    anime = pd.read_csv(f'{name_folder}\load_CPU_users.csv', encoding='windows-1251', sep=';', header=None)

    if name_user not in str(anime[2]):


        message = f'Загрузка ЦП {load}%'
        lst = [date, time, name_user, load, message]
        with open(f'{name_folder}\load_CPU_users.csv', 'a+', newline='') as file:
            writer = csv.writer(file, delimiter=';')
            writer.writerow(lst)








def parse(row_data,name_folder, folder_path_log):
    ch = r'(\)'.split('(')[1].split(')')[0]
    row_data = row_data.replace(ch, '!')
    pattern = r"(\d{2}/\d{2}/\d{4})\s(\d{2}:\d{2}:\d{2})\|\s+(\S+)\s\:\s+\d+\,\s+(\d+)\s\|\s+\d+\s+\S\s+\d+\,\s+\d+\S\,\s+(\d+)\,\s+\d+\,\s+(\d+)\,\s+(\d+)\|"
    line = re.findall(pattern, row_data)
    if line != []:
        #print(line) # [('01/08/2022', '17:03:13', '1@ua7dv.trueconf.name/456764466', '70', '65535', '500', '40')]
        date, time, ID_user, load, pkt, users_bytes = line[0][0], line[0][1], line[0][2], line[0][3], line[0][5], line[0][6]
        users_bytes = int(users_bytes)
        load = int(load)

        if users_bytes > User_byte:
            if not os.path.isdir(name_folder):
                os.mkdir(name_folder)
            create_user_csv(ID_user, name_folder, folder_path_log)

        if load > CPU_load:
            if not os.path.isdir(name_folder):
                os.mkdir(name_folder)
            load_users_csv(ID_user, name_folder, date, time, load)

    # -----------------------------------------------------------------
    #                    --Исходящий поток--
    # -----------------------------------------------------------------

def create_user_csv_sender_stat(ID_user, name_folder, folder_path_log):
    ''' Функция создания файла с исходящим потоком пользователя '''

    if '/' in ID_user:
        name_user = ID_user.replace(':', '').split('/')[0].split('@')[0]

    elif '!' in ID_user:
        name_user = ID_user.replace(':', '').split('!')[1]
        name_user = name_user.split('@')[0] if '@' in name_user else name_user

    else:
        name_user = ID_user.replace(':', '').split('@')[0]

    if os.path.exists(f'{name_folder}/{name_user}_sender_stat.xlsx') != True:
        with open(f'{name_folder}\{name_user}_sender_stat.csv', 'w', newline='', encoding='windows-1251') as file:
            header = [(
                'Дата',
                'Время',
                'ID пользователя',
                'b',
                'ba',
                'bd',
                'fv',
                'fa',
                'iv',
                'ia'
            )]
            writer = csv.writer(file, delimiter=';')
            writer.writerows(header)

            with open(f'{folder_path_log}',
                      'r') as f:  # r'C:\Users\marinich\Desktop\vs_stat_svc_00000065214657dc@ua7dv.trueconf.name#vcs.txt'
                while True:
                    row_data = f.readline()
                    ch = r'(\)'.split('(')[1].split(')')[0]
                    row_data = row_data.replace(ch, '!')
                    if not row_data:
                        break

                    patter = r"(\d{2}/\d{2}/\d{4})\s(\d{2}:\d{2}:\d{2})\|\s+"f'({ID_user})'r"\s\|\s+(\d+)\s+(\d+)\s+(\d+\s+\S+)\s+\|\s+(\d+)\s+(\d+)\s+\|\s+(\d+)\s+(\d+)"
                    line_user = re.findall(patter, row_data)
                    if line_user != []:
                        date, time, b, ba, bd, fv, fa, iv, ia = line_user[0][0], line_user[0][1], line_user[0][3], \
                                                                line_user[0][4], line_user[0][5], line_user[0][6], \
                                                                line_user[0][7], line_user[0][8], line_user[0][9]

                        lst = [date, time, name_user, b, ba, bd, fv, fa, iv, ia]

                        writer = csv.writer(file, delimiter=';')  # ,lineterminator="\r"
                        writer.writerow(lst)

        # -------------- конвертация файла с csv формата в xlsx --------------------------------------
        read_file = pd.read_csv(f'{name_folder}\{name_user}_sender_stat.csv',sep=';', encoding='windows-1251')  # считываем содержимое csv файла
        read_file.to_excel(f'{name_folder}\{name_user}_sender_stat.xlsx', index=None,header=True)  # записываем содержимое в xlsx файл
        os.remove(f'{name_folder}\{name_user}_sender_stat.csv')  # удаляем csv файл
        # --------------------------------------------------------------------------------------------

        book = openpyxl.load_workbook(f'{name_folder}/{name_user}_sender_stat.xlsx')
        sheet = book.active
        sheet.column_dimensions["A"].width = 12
        sheet.column_dimensions["B"].width = 12
        sheet.column_dimensions["C"].width = 30
        fill1 = PatternFill('solid', fgColor='00cccc')
        sheet["A1"].fill = fill1
        sheet["B1"].fill = fill1
        sheet["C1"].fill = fill1
        sheet["D1"].fill = fill1
        sheet["E1"].fill = fill1
        sheet["F1"].fill = fill1
        sheet["G1"].fill = fill1
        sheet["H1"].fill = fill1
        sheet["I1"].fill = fill1
        sheet["J1"].fill = fill1

        nb_row = sheet.max_row  # узнаём колиечтво строк в эксель файле
        time = Reference(sheet, min_col=2, min_row=2, max_row=nb_row)


        book.create_sheet("Chart")
        sheet2 = book["Chart"]

        chart3 = LineChart()
        chart3.title = "iv (Максимальный интервал времени между пакетам видео)"
        chart3.y_axis.title = "мс"
                                                    # График 1
        chart3.x_axis.number_format = 'H-M'
        chart3.x_axis.majorTimeUnit = "time"
        chart3.x_axis.title = "Время"

        chart3.add_data(f'Sheet1!I2:I{nb_row}')
        chart3.set_categories(time)
        chart3.width = 30
        chart3.height = 10

        sheet2.add_chart(chart3, "A1")
        #---------------------------------------------
        chart4 = LineChart()
        chart4.title = "ia (Максимальный интервал времени между пакетам аудио)"
        chart4.y_axis.title = "мс"
                                                    # График 2
        chart4.x_axis.number_format = 'H-M'
        chart4.x_axis.majorTimeUnit = "time"
        chart4.x_axis.title = "Время"

        chart4.add_data(f'Sheet1!J2:J{nb_row}')
        chart4.set_categories(time)
        chart4.width = 30
        chart4.height = 10

        sheet2.add_chart(chart4, "A22")
        #---------------------------------------------
        chart = LineChart()
        chart.title = "fv (Отсылаемая частота кадров видео)"
        chart.y_axis.title = "FPS"
                                                    # График 3
        chart.x_axis.number_format = 'H-M'
        chart.x_axis.majorTimeUnit = "time"
        chart.x_axis.title = "Время"

        chart.add_data(f'Sheet1!G2:G{nb_row}')
        chart.set_categories(time)
        chart.width = 30
        chart.height = 10

        sheet2.add_chart(chart, "A44")
        #---------------------------------------------
        chart2 = LineChart()
        chart2.title = "fa (Отсылаемая частота пакетов аудио)"
        chart2.y_axis.title = "Частота"
                                                    # График 4
        chart2.x_axis.number_format = 'H-M'
        chart2.x_axis.majorTimeUnit = "time"
        chart2.x_axis.title = "Время"

        chart2.add_data(f'Sheet1!H2:H{nb_row}')
        chart2.set_categories(time)
        chart2.width = 30
        chart2.height = 10

        sheet2.add_chart(chart2, "A66")
        #---------------------------------------------

        table = Table(displayName="Sheet1", ref=f'A1:J{nb_row}')  # создаем объект таблицы
        sheet.add_table(table)  # добавляем таблицу

        book.save(f'{name_folder}/{name_user}_sender_stat.xlsx')  # сохраняем файл

#------------------------------------------------------------------------------



class Window(QMainWindow):
    def __init__(self):
        super(Window, self).__init__()


        self.setWindowTitle('Парсер svc_logs')
        self.setGeometry(300, 200, 700, 400)

        self.createMenuBar()

        self.text_edit = QtWidgets.QTextEdit(self)
        self.text_edit.setReadOnly(True)   # запрещает пользователю редактировать текст


        self.setCentralWidget(self.text_edit)
        self.button = QtWidgets.QPushButton(self)
        self.setCentralWidget(self.text_edit)
        grid = QtWidgets.QGridLayout(self.text_edit)
        #self.button.move(250, 350)
        self.button.setText('Анализ')
        self.button.setFixedWidth(200)  # укзываем ширину кнопки
        self.button.setFixedHeight(40)  # указываем высоту кнопки
        self.button.setEnabled(False)
        self.button.clicked.connect(self.flow1)
        grid.addWidget(self.button, 0, 0, QtCore.Qt.AlignBaseline | QtCore.Qt.AlignBottom)





    def start_analysis(self):

        self.text_edit.append('Начало выполнения скрипта...\n\n')
        self.button.setEnabled(False)
        print('Начало выполнения скрипта...')
        try:
            start_time = time.time()
            i = 0
            while i < len(self.fnames):
                with open(self.fnames[i], 'r') as f:
                    name_folder = '.'.join(self.fnames[i].split('/')[-1].split('.')[:-1])
                    #print('Name_folder:', name_folder) # vs_stat_svc_000000a06aab935d@ua7dv.trueconf.name#vcs
                    folder_path_log = self.fnames[i]
                    #print('Folder_path_log:', folder_path_log) # C:/TrueConf/svc_logs/vs_stat_svc_000000a06aab935d@ua7dv.trueconf.name#vcs.txt
                    while True:
                        row_data = f.readline()

                        if not row_data:
                            break
                        if re.sub(r"\s+$", "", row_data).endswith('|'):
                            parse(row_data, name_folder, folder_path_log)

                    if os.path.isdir(name_folder):
                        if os.path.exists(f'{name_folder}\load_CPU_users.csv') == True:

                            # -------------- конвертация файла load_CPU_users.csv с csv формата в xlsx  --------------------------------------
                            read_file = pd.read_csv(f'{name_folder}\load_CPU_users.csv', sep=';', encoding='windows-1251')  # считываем содержимое csv файла
                            read_file.to_excel(f'{name_folder}\load_CPU_users.xlsx', index=None, header=True)  # записываем содержимое в xlsx файл
                            os.remove(f'{name_folder}\load_CPU_users.csv')  # удаляем csv файл
                                        #----- Стиль файла load_CPU_users.xlsx ------

                            book = openpyxl.load_workbook(f'{name_folder}/load_CPU_users.xlsx')
                            sheet = book.active
                            sheet.column_dimensions["A"].width = 12
                            sheet.column_dimensions["B"].width = 12
                            sheet.column_dimensions["C"].width = 30
                            sheet.column_dimensions["D"].width = 17
                            sheet.column_dimensions["E"].width = 20
                            fill1 = PatternFill('solid', fgColor='00cccc')
                            sheet["A1"].fill = fill1
                            sheet["B1"].fill = fill1
                            sheet["C1"].fill = fill1
                            sheet["D1"].fill = fill1
                            sheet["E1"].fill = fill1

                            nb_row = sheet.max_row  # узнаём колиечтво строк в эксель файле
                            print(nb_row)
                            table = Table(displayName="Sheet1", ref=f'A1:E{nb_row}')  # создаем объект таблицы
                            sheet.add_table(table)  # добавляем таблицу

                            book.save(f'{name_folder}/load_CPU_users.xlsx')
                            # --------------------------------------------------------------------------------------------

                        files = os.listdir(path=f"{name_folder}")
                        while i < len(files):
                            if files[i] != f'{name_a_shared_csv}.csv' and files[i] != 'load_CPU_users.xlsx' and files[i] != 'load_CPU_users.csv' and ('sender_stat' not in files[i]):
                                try:
                                    creating_a_shared_csv(name_folder)
                                except ValueError:
                                    self.text_edit.append('Обнаружена ошибка в модуле shared, работа скрипта продолжается \n')
                                break
                            i += 1

                    i += 1

            self.text_edit.append('\n')
            self.text_edit.append('Конец выполнения...')
            print('Конец выполнения ...')
            print("--- %s seconds ---" % (time.time() - start_time))
            self.text_edit.append("--- %s seconds ---" % round((time.time() - start_time), 3))


        except FileNotFoundError:
            print('No such File')
        self.button.setEnabled(False)
#-----------------------------------------------------------------
#                    --Исходящий поток--
#-----------------------------------------------------------------


    def start_analusis_sender_stat(self):
        try:
            it = 0
            while it < len(self.fnames):
                with open(self.fnames[it], 'r') as f:
                    name_folder = '.'.join(self.fnames[it].split('/')[-1].split('.')[:-1])
                    # print('Name_folder:', name_folder) # vs_stat_svc_000000a06aab935d@ua7dv.trueconf.name#vcs
                    folder_path_log = self.fnames[it]
                    # print('Folder_path_log:', folder_path_log) # C:/TrueConf/svc_logs/vs_stat_svc_000000a06aab935d@ua7dv.trueconf.name#vcs.txt
                    while True:
                        row_data = f.readline()

                        if not row_data:
                            break

                        ch = r'(\)'.split('(')[1].split(')')[0]
                        row_data = row_data.replace(ch, '!')
                        pattern = r"(\d{2}/\d{2}/\d{4})\s(\d{2}:\d{2}:\d{2})\|\s+(\S+)\s\|\s+(\d+)\s+(\d+)\s+(\d+\s+\S+)\s+\|\s+(\d+)\s+(\d+)\s+\|\s+(\d+)\s+(\d+)"
                        line = re.findall(pattern, row_data)
                        if line != []:
                            date, time, ID_user, b, ba, bd, fv, fa, iv, ia = line[0][0], line[0][1], line[0][2], \
                                                                             line[0][3], line[0][4], line[0][5], \
                                                                             line[0][6], line[0][7], line[0][8], \
                                                                             line[0][9]
                            fv, fa, iv, ia = int(fv), int(fa), int(iv), int(ia)

                            if fv > fv_input or fa > fa_input or iv > iv_input or ia > ia_input:  # если какое-то значние больше, то считаем строку проблемной
                                if not os.path.isdir(name_folder):
                                    os.mkdir(name_folder)
                                create_user_csv_sender_stat(ID_user, name_folder, folder_path_log)

                    it += 1

        except FileNotFoundError:
            print('No such File Sender Stat')
        self.button.setEnabled(False)


#--------------------------------------------------------------------


# ---------------ПОТОКИ------------------
    def flow1(self):  # метод вызова функции start_analysis в отдельном потоке
        t = threading.Thread(target=self.start_analysis) # поток для входящего битрейта
        t.start()


        if int(flag_sender_stat) == 1:
            t2 = threading.Thread(target=self.start_analusis_sender_stat) # поток для исходящего битрейта
            t2.start()
# --------------------------------------

    def createMenuBar(self):
        self.menuBar = QMenuBar(self)
        self.setMenuBar(self.menuBar)

        fileMenu = QMenu('&Файл', self)
        self.menuBar.addMenu(fileMenu)

        fileMenu.addAction('Открыть', self.action_clicked)
        fileMenu.addSeparator()
        fileMenu.addAction('Выход из программы', self.action_clicked)



    @QtCore.pyqtSlot()
    def action_clicked(self):

        action = self.sender()
        if action.text() == 'Открыть':
            print('Open')
            self.fnames = QFileDialog.getOpenFileNames(self)[0]   # QFileDialog.getOpenFileNames(self)[0] для открытия нескольких файлов

            if self.fnames != None and self.fnames != [] :
                self.button.setEnabled(True)

            self.text_edit.clear()

        elif action.text() == 'Выход из программы':
            sys.exit(app.exec_())




if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = Window()
    window.show()
    sys.exit(app.exec_())
